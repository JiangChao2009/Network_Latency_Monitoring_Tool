__author__ = 'Gareth'

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import cell

import os.path
import subprocess
import time
import datetime

# SETTINGS~~~~~~~~~~~~~~~~~
server = "www.google.com"  # server to ping
time_between_pings = 10  # time between pings (s)
time_between_saves = 30  # time between saves (s)
# ~~~~~~~~~~~~~~~~~~~~~~~~

if os.path.exists('Readings.xlsx') is False:
    wb = Workbook()
    ws = wb.active

    ws['A1'] = "Number of Readings:"
    ws['B1'] = 0
    ws['A2'] = "Latency (ms)"
    ws['B2'] = "Time in Seconds from Start of Day"
    ws['C2'] = "Date/Time"
    ws['D2'] = "Date"

    wb.save('Readings.xlsx')

else:
    wb = load_workbook('Readings.xlsx')
    ws = wb.active

time_since_last_save = 0
save_time = 0

while 1 == 1:
    # pings a server and returns the ping time (in ms)
    ping = subprocess.Popen(["ping.exe", server], stdout=subprocess.PIPE)
    ping2 = ping.communicate()[0]
    ping2 = str(ping2)
    ping2 = ping2.split('time=', 1)[1]
    ping2 = ping2.split('ms', 1)[0]

    # takes the present time, and then returns the time from the start of the day in seconds
    present_time = datetime.datetime.now()
    present_time_total_seconds = (present_time.hour*3600) + (present_time.minute*60) + present_time.second

    row_num = ws['B1'].value + 3

    ws.cell(row=row_num, column=1).value = int(ping2)
    ws.cell(row=row_num, column=2).value = int(present_time_total_seconds)
    ws.cell(row=row_num, column=3).value = str(present_time)
    ws.cell(row=row_num, column=4).value = str(present_time.date())

    ws['B1'].value += 1
    time_since_last_save = present_time_total_seconds - save_time

    if time_since_last_save > time_between_saves:
        wb.save("Readings.xlsx")
        save_time = present_time_total_seconds
        print("SAVED TO EXCEL")

    print("Time:%s Latency:%dms" % (str(present_time), int(ping2)))
    time.sleep(time_between_pings)
